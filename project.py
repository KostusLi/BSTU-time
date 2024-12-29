import json
from openpyxl import load_workbook

def read_schedule_with_subgroups_from_excel(excel_path, json_path):
    # Открываем Excel-документ
    workbook = load_workbook(excel_path)
    sheet = workbook.active  # Используем активный лист

    # Создаем словарь для хранения расписания
    schedule = {}

    # Список возможных дней недели
    days_of_week = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

    # Считаем, что первая строка содержит заголовки
    headers = [cell.value.strip() if cell.value else "" for cell in sheet[1]]

    # Проверяем, есть ли заголовки
    if not headers:
        raise ValueError("Excel файл не содержит заголовков в первой строке.")

    # Определяем индексы для колонок "Дни", "Часы" и групп
    try:
        day_index = headers.index("Дни")
        time_index = headers.index("Часы")
    except ValueError:
        raise ValueError("Отсутствуют колонки с заголовками 'Дни' и 'Часы'.")

    # Получаем индексы групп (остальные колонки)
    group_indices = [i for i, header in enumerate(headers) if i != day_index and i != time_index]

    # Переменные для хранения текущих дней и времени
    current_day = None
    current_time = None

    # Считываем строки таблицы, начиная со второй (индекс 2)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cells = [str(cell).strip() if cell else "" for cell in row]

        # Обновляем день недели, если ячейка не пуста
        day = cells[day_index] if day_index < len(cells) and cells[day_index] else current_day
        if day in days_of_week:
            current_day = day

        # Обновляем время только если ячейка содержит новый интервал
        if time_index < len(cells) and cells[time_index]:
            current_time = cells[time_index]

        # Если нет текущего дня или времени, пропускаем строку
        if not current_day or not current_time:
            continue

        # Обрабатываем данные для каждой группы
        for group_index in group_indices:
            if group_index < len(cells):
                group_name = headers[group_index]
                lesson_details = cells[group_index]

                if not lesson_details:
                    continue

                # Проверяем, разделена ли колонка на две подгруппы (по разделителю '|')
                if "|" in lesson_details:
                    left_part, right_part = map(str.strip, lesson_details.split("|", 1))

                    # Первая подгруппа
                    subgroup1 = f"{group_name.strip()} подгруппа 1"
                    if subgroup1 not in schedule:
                        schedule[subgroup1] = {}

                    if current_day not in schedule[subgroup1]:
                        schedule[subgroup1][current_day] = []

                    existing_entries_1 = schedule[subgroup1][current_day]
                    if not any(entry["Часы"] == current_time and entry["Занятия"] == left_part for entry in existing_entries_1):
                        schedule[subgroup1][current_day].append({"Часы": current_time, "Занятия": left_part})

                    # Вторая подгруппа
                    subgroup2 = f"{group_name.strip()} подгруппа 2"
                    if subgroup2 not in schedule:
                        schedule[subgroup2] = {}

                    if current_day not in schedule[subgroup2]:
                        schedule[subgroup2][current_day] = []

                    existing_entries_2 = schedule[subgroup2][current_day]
                    if not any(entry["Часы"] == current_time and entry["Занятия"] == right_part for entry in existing_entries_2):
                        schedule[subgroup2][current_day].append({"Часы": current_time, "Занятия": right_part})

                else:  # Если колонка объединена и не содержит разделителя
                    subgroup1 = f"{group_name.strip()} подгруппа 1"
                    subgroup2 = f"{group_name.strip()} подгруппа 2"

                    for subgroup in [subgroup1, subgroup2]:
                        if subgroup not in schedule:
                            schedule[subgroup] = {}

                        if current_day not in schedule[subgroup]:
                            schedule[subgroup][current_day] = []

                        existing_entries = schedule[subgroup][current_day]
                        if not any(entry["Часы"] == current_time and entry["Занятия"] == lesson_details for entry in existing_entries):
                            schedule[subgroup][current_day].append({"Часы": current_time, "Занятия": lesson_details})

    # Сохраняем данные в JSON файл
    with open(json_path, 'w', encoding='utf-8') as json_file:
        json.dump(schedule, json_file, ensure_ascii=False, indent=4)

# Укажите путь к вашему Excel-документу и выходному файлу JSON
excel_path = 'myfile.xlsx'  # Замените на ваш путь
json_path = 'table.json'

# Вызов функции
read_schedule_with_subgroups_from_excel(excel_path, json_path)